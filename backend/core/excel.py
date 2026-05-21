import csv
import os
import io
import openpyxl
from collections import Counter
from datetime import datetime, date
from typing import Any
from backend.utils.helpers import debug_print


def _fix_str(value):
    if not isinstance(value, str):
        return value
    try:
        return value.encode("latin-1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return value


def _csv_encoding(path):
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with open(path, encoding=enc) as f:
                f.read(100)
            return enc
        except (UnicodeDecodeError, UnicodeError):
            continue
    return "utf-8"


def get_sheet_names(file_path: str) -> list[dict[str, Any]]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        encoding = _csv_encoding(file_path)
        with open(file_path, encoding=encoding) as f:
            rows = sum(1 for _ in f)
        name = os.path.splitext(os.path.basename(file_path))[0]
        return [{"name": name, "rows": rows}]
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheets = [{"name": name, "rows": wb[name].max_row} for name in wb.sheetnames]
    wb.close()
    return sheets


def parse_sheet(file_path: str, sheet_name: str) -> dict[str, Any]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return _parse_csv(file_path)
    return _parse_excel(file_path, sheet_name)


def _detect_delimiter(content: str) -> str:
    try:
        sample = content[:4096]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        debug_print(f"[CSV] detected delimiter: {repr(dialect.delimiter)}")
        return dialect.delimiter
    except csv.Error:
        debug_print("[CSV] could not detect delimiter, defaulting to ','")
        return ","


def _parse_csv(file_path: str) -> dict[str, Any]:
    encoding = _csv_encoding(file_path)
    debug_print(f"[CSV] encoding: {encoding}")
    with open(file_path, encoding=encoding) as f:
        content = f.read()
    content = _fix_str(content)
    delimiter = _detect_delimiter(content)
    reader = csv.DictReader(io.StringIO(content), delimiter=delimiter)
    headers = [_fix_str(h) for h in reader.fieldnames]
    debug_print(f"[CSV] headers: {headers}")
    rows = []
    for row in reader:
        fixed = {}
        for k, v in row.items():
            fixed[_fix_str(k)] = _fix_str(v) if isinstance(v, str) else v
        rows.append(fixed)
    debug_print(f"[CSV] parsed rows: {len(rows)}")
    return {"headers": headers, "rows": rows, "total_rows": len(rows)}


def _parse_excel(file_path: str, sheet_name: str) -> dict[str, Any]:
    debug_print(f"[EXCEL] sheet: {sheet_name}")
    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    headers = [_fix_str(cell.value) for cell in next(ws.iter_rows(max_row=1))]
    debug_print(f"[EXCEL] headers: {headers}")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = [_fix_str(v) if isinstance(v, str) else v for v in row]
        rows.append(dict(zip(headers, vals)))
    wb.close()
    debug_print(f"[EXCEL] parsed rows: {len(rows)}")
    return {"headers": headers, "rows": rows, "total_rows": len(rows)}


def _detect_type(values: list) -> str:
    non_null = [v for v in values if v is not None]
    if not non_null:
        return "text"
    numeric = sum(1 for v in non_null if isinstance(v, (int, float)))
    if numeric / len(non_null) > 0.8:
        return "numeric"
    dates = sum(1 for v in non_null if isinstance(v, (datetime, date)))
    if dates / len(non_null) > 0.8:
        return "date"
    try:
        parsed = 0
        for v in non_null:
            float(v)
            parsed += 1
        if parsed / len(non_null) > 0.8:
            return "numeric"
    except (ValueError, TypeError):
        pass
    return "text"


_MONETARY_KEYS = {"monto", "valor", "precio", "costo", "total", "importe", "amount", "price", "cost", "saldo", "balance", "ingreso", "egreso"}
_CATEGORY_KEYS = {"categoria", "categoría", "category", "tipo", "type", "rubro", "rubro", "clase", "class"}
_DATE_KEYS = {"fecha", "date", "fecha"}


def compute_summary(headers: list[str], rows: list[dict]) -> dict[str, Any]:
    summary = {"total_rows": len(rows), "columns": [], "hints": {}}
    for header in headers:
        col_vals = [r.get(header) for r in rows]
        col_type = _detect_type(col_vals)
        non_null = [v for v in col_vals if v is not None]
        col_info = {"name": header, "type": col_type, "non_null": len(non_null), "nulls": len(rows) - len(non_null)}
        if col_type == "numeric":
            nums = []
            for v in col_vals:
                try:
                    nums.append(float(v))
                except (ValueError, TypeError):
                    pass
            if nums:
                col_info["min"] = round(min(nums), 2)
                col_info["max"] = round(max(nums), 2)
                col_info["sum"] = round(sum(nums), 2)
                col_info["avg"] = round(sum(nums) / len(nums), 2)
        elif col_type == "text":
            counts = Counter(str(v) for v in non_null)
            col_info["top"] = counts.most_common(10)
        summary["columns"].append(col_info)
        hl = header.lower().strip()
        if hl in _MONETARY_KEYS or any(k in hl for k in _MONETARY_KEYS):
            summary["hints"][header] = "monetary"
        elif hl in _CATEGORY_KEYS or any(k in hl for k in _CATEGORY_KEYS):
            summary["hints"][header] = "category"
        elif hl in _DATE_KEYS or any(k in hl for k in _DATE_KEYS):
            summary["hints"][header] = "date"
    return summary


def get_sample(rows: list[dict], n: int = 10) -> list[dict]:
    return rows[:n]