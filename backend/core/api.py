import sqlite3
import json
import os
import openpyxl
import webview
import markdown
from datetime import datetime
from typing import Any
from backend.core.llm import LLMClient

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "database", "history.db")


def _init_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS reports (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            currency TEXT NOT NULL,
            language TEXT NOT NULL,
            raw_md TEXT NOT NULL,
            html TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    """)
    conn.commit()
    conn.close()


def _md_to_html(md_text: str) -> str:
    return markdown.markdown(md_text, extensions=["nl2br", "sane_lists"])


class API:
    def __init__(self, llm_client: LLMClient):
        self._llm = llm_client
        _init_db()

    def minimize_window(self) -> None:
        if webview.windows:
            webview.windows[0].minimize()

    def close_window(self) -> None:
        if webview.windows:
            webview.windows[0].destroy()

    def open_file_dialog(self) -> dict[str, Any]:
        try:
            result = webview.windows[0].create_file_dialog(
                webview.FileDialog.OPEN,
                allow_multiple=False,
                file_types=("Excel Files (*.xlsx;*.xls)",),
            )
            if result and len(result) > 0:
                return {"success": True, "path": result[0]}
            return {"success": False, "error": "No file selected"}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def get_model_info(self) -> dict[str, str]:
        model = self._llm.ollama_model if self._llm.use_ollama else self._llm.nvidia_model
        return {"model": model}

    def get_provider(self) -> dict[str, Any]:
        return {"ollama": self._llm.use_ollama, "provider": "ollama" if self._llm.use_ollama else "nvidia"}

    def process_excel(self, file_path: str) -> dict[str, Any]:
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
            filename = os.path.basename(file_path)
            return {"success": True, "data": rows, "filename": filename}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def generate_report(self, json_data: str, filename: str, currency: str = "original", language: str = "Español") -> dict[str, Any]:
        try:
            raw_md = self._llm.generate(json_data, currency=currency, language=language)
            if raw_md is None:
                return {"success": False, "error": "No response from LLM"}
            html = _md_to_html(raw_md)
            conn = sqlite3.connect(DB_PATH)
            conn.execute(
                "INSERT INTO reports (filename, currency, language, raw_md, html, created_at) VALUES (?, ?, ?, ?, ?, ?)",
                (filename, currency, language, raw_md, html, datetime.now().isoformat())
            )
            conn.commit()
            conn.close()
            return {"success": True, "html": html, "raw_md": raw_md}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def get_history(self) -> dict[str, Any]:
        try:
            conn = sqlite3.connect(DB_PATH)
            rows = conn.execute(
                "SELECT id, filename, currency, language, created_at FROM reports ORDER BY created_at DESC LIMIT 50"
            ).fetchall()
            conn.close()
            items = [
                {"id": r[0], "filename": r[1], "currency": r[2], "language": r[3], "created_at": r[4]}
                for r in rows
            ]
            return {"success": True, "items": items}
        except Exception as e:
            return {"success": False, "error": str(e)}

    def get_report(self, report_id: int) -> dict[str, Any]:
        try:
            conn = sqlite3.connect(DB_PATH)
            row = conn.execute(
                "SELECT id, filename, currency, language, raw_md, html, created_at FROM reports WHERE id = ?",
                (report_id,)
            ).fetchone()
            conn.close()
            if not row:
                return {"success": False, "error": "Report not found"}
            return {
                "success": True,
                "id": row[0], "filename": row[1], "currency": row[2],
                "language": row[3], "raw_md": row[4], "html": row[5], "created_at": row[6]
            }
        except Exception as e:
            return {"success": False, "error": str(e)}

    def delete_report(self, report_id: int) -> dict[str, Any]:
        try:
            conn = sqlite3.connect(DB_PATH)
            conn.execute("DELETE FROM reports WHERE id = ?", (report_id,))
            conn.commit()
            conn.close()
            return {"success": True}
        except Exception as e:
            return {"success": False, "error": str(e)}